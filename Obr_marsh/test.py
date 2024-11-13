

def change_table_2(table_2):
    ''' Редактирует таблицу 2 (замена . на , и добавление хвостовых нулей'''
    for i in table_2:
        for k in i.keys():
            i[k] = i[k].replace('.', ',')
            if len(i[k]) == 1:
                i[k] += ',0'
        i['km_nach'] += '00'
        i['km_kon'] += '00'


def change_table_3(table_3):
    ''' Редактирует таблицу 3 (замена . на , и добавление хвостовых нулей'''
    for i in table_3:
        for k in i.keys():
            i[k] = i[k].replace('.', ',')
            if len(i[k]) == 1:
                i[k] += ',0'
            if len(i['kpr_i']) == 3:
                i['kpr_i'] += '0'

def change_table_4(table_4):
    ''' Редактирует таблицу 4 (замена . на , и добавление хвостовых нулей'''
    for i in table_4:
        for k in i.keys():
            i[k] = i[k].replace('.', ',')
            if len(i['kpr_i']) == 3:
                i['kpr_i'] += '0'


import openpyxl as xl
from docxtpl import DocxTemplate
from docxcompose.composer import Composer

import os
import docx

from start_table import start_table
from WGS_Table import wgs_table
from some_tip import some_tips
from osnov_vid_def import vivodi_v_otchet


workbook = xl.load_workbook('Ведомость.xlsx', data_only=True)
sheet_names = workbook.sheetnames

if __name__ == '__main__':
    workbook = xl.load_workbook('Ведомость.xlsx', data_only=True)
    sheet_names = [i for i in workbook.sheetnames if i not in ['Лист1', 'ИД', 'В обсл', 'аб1']]
    print(sheet_names)




